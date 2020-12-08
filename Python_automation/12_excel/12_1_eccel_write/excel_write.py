import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet_name = wb.sheetnames
sheet = wb[sheet_name[0]]

price_update = {'Celery': 1.19,
                'Garlic': 3.07,
                'Lemon': 1.27}

for row in range(2, sheet.max_row+1):
    produce_name = sheet.cell(row=row, column=1).value
    if produce_name in price_update:
        sheet.cell(row=row, column=2).value = price_update[produce_name]

wb.save('updateproduceSales.xlsx')