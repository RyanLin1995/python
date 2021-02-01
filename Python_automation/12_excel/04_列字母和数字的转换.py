import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook("/home/ryan/Downloads/automate_online-materials/example.xlsx")
sheet = wb.active

print(get_column_letter(1))
print(column_index_from_string("A"))
print(f"Max column: {sheet.max_column}")
print(get_column_letter(sheet.max_column))
