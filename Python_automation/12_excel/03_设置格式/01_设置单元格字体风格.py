from openpyxl.styles import Font
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test"
ws["A1"] = "Hello World"
sheet_font = Font(size=24, italic=True)
ws["A1"].font = sheet_font
wb.save("01text.xlsx")