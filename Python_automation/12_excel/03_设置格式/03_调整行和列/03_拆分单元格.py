import openpyxl

wb = openpyxl.load_workbook("02test.xlsx")
ws = wb.active
ws.unmerge_cells("A1:D3")
ws.unmerge_cells("C5:D5")
wb.save("02test.xlsx")