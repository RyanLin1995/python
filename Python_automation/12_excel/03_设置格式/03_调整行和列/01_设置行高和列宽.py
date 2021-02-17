import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "Tall Row"
ws["B2"] = "Wide Column"
ws.row_dimensions[1].height = 70  # 设置行高，当数值为0，则隐藏
ws.column_dimensions["B"].height = 20  # 设置列宽，当数值为0，则隐藏
wb.save("01test.xlsx")