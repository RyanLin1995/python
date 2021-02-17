import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.merge_cells("A1:D3")  # 合并单元格，参数填写的是需要合并的区域
ws["A1"] = "This is a merge cell"  # 如果要写入合并单元格数值，只需设置第一个区域的值即可
ws.merge_cells("C5:D5")
ws["C5"] = "This is the other merge cell"
wb.save("02test.xlsx")