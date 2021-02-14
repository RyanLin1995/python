import openpyxl

# 写入公式
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = 100
ws["A2"] = 200
ws["A3"] = "=SUM(A1:A2)"
wb.save("02test.xlsx")

# 读取带公式的表格
# 读取带公式的表格时，分两种情况
# 情况一：只读取数值(Windows下有用，需要先打开目标文件，保存一次)
wb = openpyxl.load_workbook("02test.xlsx", data_only=True)
ws = wb.active
print(ws["A3"].value)  # 要先打开一次目标文件保存后，才会有数值。因为目前没有缓存，所以读出来是None(只在Windows下生效)

# 情况二：直接读取单元格内容
wb = openpyxl.load_workbook("02test.xlsx")
ws = wb.active
print(ws["A3"].value)