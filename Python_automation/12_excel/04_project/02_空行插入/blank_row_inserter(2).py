import openpyxl

info = input("请输入信息(需要插入的行号， 需要插入的数量， 文件名称): ")
info_update = info.split(",")
if len(info_update) < 3 or len(info_update) > 4:
    print("输入参数有无，最多只接受三个参数")

wb = openpyxl.load_workbook(info_update[-1])
ws = wb.active

wb1 = openpyxl.Workbook()
ws1 = wb1.active

for i in range(1, int(info_update[0])):
    for j in range(1, ws.max_column + 1):
        ws1.cell(row=i, column=j).value = ws.cell(row=i, column=j).value

for i in range(int(info_update[0]), ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        ws1.cell(row=i + int(info_update[1]), column=j).value = ws.cell(row=i, column=j).value

wb1.save(f"(new2){info_update[-1]}")

