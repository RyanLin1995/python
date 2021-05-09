import datetime
import openpyxl

wb = openpyxl.load_workbook("1.xlsx")
ws = wb.active

wb1 = openpyxl.Workbook()
ws1 = wb1.active


def check_time(func):
    def check():
        start_time = datetime.datetime.now()
        func()
        end_time = datetime.datetime.now()
        print(f"time: {end_time - start_time}")
    return check


def read_data():

    for wb_row in range(0, ws.max_row):
        for wb_column in range(0, ws.max_column):
            yield ws.cell(row=wb_row + 1, column=wb_column + 1).value


@check_time
def write_data():

    data = read_data()
    for wb_column in range(1, 16385):
        for wb_row in range(1, ws.max_column + 1):
            ws1.cell(row=wb_row, column=wb_column).value = next(data)

    wb1.save("2.xlsx")


if __name__ == '__main__':

    write_data()

