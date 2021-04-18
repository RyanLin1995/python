import openpyxl

num = input("输入需要创建的乘法表的号码：")
if int(num) > 9 or int(num) <= 0:
    print("无效的乘法表数值")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Multiplication Table"

for i in range(1, int(num) + 1):
    ws.cell(row=(i+1), column=1).value = i
    ws.cell(row=1, column=(i+1)).value = i

for j in range(1, int(num) + 1):
    for k in range(2, int(num) + 2):
        ws.cell(row=k, column=j+1).value = ws.cell(row=1, column=k).value * ws.cell(row=j+1, column=1).value

wb.save(r"multiplication_table.xlsx")
