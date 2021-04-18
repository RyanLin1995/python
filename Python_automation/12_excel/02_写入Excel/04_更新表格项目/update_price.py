import openpyxl

wb = openpyxl.load_workbook("/home/ryan/Downloads/automate_online-materials/produceSales.xlsx")
ws = wb.active

# The produce types and their updated prices
update_price = {
    "Celery": 1.19,
    "Garlic": 3.07,
    "Lemon": 1.27
}

# Todo Loop through the rows and update the prices
for row_num in range(2, ws.max_row + 1):
    produce_name = ws.cell(row=row_num, column=1).value
    if produce_name in update_price:
        ws.cell(row=row_num, column=2).value = update_price[produce_name]

wb.save("(new)1.xlsx")