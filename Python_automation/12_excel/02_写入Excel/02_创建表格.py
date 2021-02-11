import openpyxl

wb = openpyxl.Workbook()
print(wb.sheetnames)

# 直接创建表格
wb.create_sheet()
print(wb.sheetnames)

# 根据索引创建表格
wb.create_sheet(index=0, title="First sheet")
print(wb.sheetnames)
wb.create_sheet(index=2, title="Middle sheet")
print(wb.sheetnames)

# 删除表格
wb.remove(wb['Sheet'])
print(wb.sheetnames)

wb.save("02_test.xlsx")

