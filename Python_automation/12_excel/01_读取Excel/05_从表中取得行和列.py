import openpyxl

wb = openpyxl.load_workbook("/home/ryan/Downloads/automate_online-materials/example.xlsx")
sheet = wb["Sheet1"]
print(tuple(sheet["A1":"C3"]))  # 从 A1 到 C3 的矩形中取出 Cell 对象，得到一个 Generator 对象。对象中包含3个元组，每个元组代表一行

for cell_obj in sheet["A1":"C3"]:
    for cell in cell_obj:
        print(cell.coordinate, cell.value)
    print("--- END OF ROWS ---")

wb.close()
print()

# 也可以使用 rows 或 columns 访问特定的行或列中所有的单元格
wb1 = openpyxl.load_workbook("/home/ryan/Downloads/automate_online-materials/example.xlsx")
sheet1 = wb1.active

print("Column")
for columns in sheet1.columns:
    for column_cell in columns:
        print(column_cell.coordinate, column_cell.value)
    print("--- END OF ROWS ---")

print()
print("Row")
for rows in sheet1.rows:
    for row_cell in rows:
        print(row_cell.coordinate, row_cell.value)
    print("--- END OF ROWS ---")

# 还可以利用 iter_rows 跟 iter_cols 指定获取数据的区域
print(tuple(sheet1.iter_rows(min_row=1, max_row=3, max_col=1)))  # 获取 column1, row1 - row3 的数据
print(tuple(sheet1.iter_cols())[sheet1.max_column - 1])  # 仅获取最大 column 的数据
