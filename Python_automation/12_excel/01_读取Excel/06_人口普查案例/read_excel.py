import openpyxl
import pprint

wb = openpyxl.load_workbook(
    "/home/ryan/Downloads/automate_online-materials/censuspopdata.xlsx")
ws = wb.active
country_data = {}

# TODO Fill in country_data with each country's population and tracts
for row in range(2, ws.max_row + 1):
    state = ws["B" + str(row)].value
    country = ws["C" + str(row)].value
    pop = ws["D" + str(row)].value

    # Make sure the key for this state exists
    country_data.setdefault(state, {})

    # Make sure the key for this country exists
    country_data[state].setdefault(country, {'tracts':0, 'pop':0})

    # Each row represents one census tract, so increment by one.
    country_data[state][country]['tracts'] += 1

    # Increase the country pop by the pop in this census tract
    country_data[state][country]['pop'] += int(pop)


# TODO Open a text file and write the contents of country data in it
result_file = open("census2010.py", "w")
result_file.write("data = " + pprint.pformat(country_data))
result_file.close()