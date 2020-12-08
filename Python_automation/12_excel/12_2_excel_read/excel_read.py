import openpyxl
import pprint

wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet_name = wb.sheetnames
sheet = wb[sheet_name[0]]

countrydata = {}
for row in range(2, sheet.max_row+1):
    state = sheet['B' + str(row)].value
    country = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    countrydata.setdefault(state, {})
    countrydata[state].setdefault(country, {'traces': 0, 'pop': 0})
    countrydata[state][country]['traces'] += 1
    countrydata[state][country]['pop'] += int(pop)

resultFile = open('census2010.py', 'w')
resultFile.write('alldata= ' + pprint.pformat(countrydata))
resultFile.close()
print('Done')

