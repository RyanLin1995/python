import openpyxl

wb = openpyxl.load_workbook(r"/home/ryan/Downloads/automate_online-materials/example.xlsx")
sheet = wb["Sheet1"]
print(sheet["A1"])  # 得到一个 Cell 对象
print(sheet["A1"].value)  # 得到 A1 的值
c = sheet["B1"]
print(c.value)
print(f"Row {c.row},Column {c.column} is {c.value}")  # 可以通过 Cell 对象的 row 跟 column 属性获取其 row 跟 column
print(f"Cell {c.coordinate} is c.value")  # c.coordinate 返回的是 单元格的位置

d = sheet.cell(row=1, column=2)  # 也可以通过给 shett.cell传递 row 跟 column 参数取值
print(d)
print(d.value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)


print(f"The highest row of Sheet1 is {sheet.max_row}")  # 获取最高行数
print(f"The highest column of Sheet1 is {sheet.max_column}")  # 获取最高列数
